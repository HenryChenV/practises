#-*- coding=utf-8 -*-

import string
from openpyxl import Workbook
from openpyxl.styles import Style, Font
import MySQLdb
from MySQLdb.cursors import DictCursor
import traceback
from datetime import datetime


MYSQL_CONF = {
    'host': 'localhost',
    'port': 3808,
    'db': 'mysite',
    'user': 'work',
    'passwd': 'xxxxxx',
}

SQL_STR = '''
    select * from polls_poll;
'''

FILE_NAME = 'test.xlsx'


class ExcelHelper(object):

    @staticmethod
    def RecordsToExcel(records, columns, out_file, sheet_name = u'AutoSheet'):
        ''' records要求是一个tuple/list，里面所有的元素是tuple/list
            column_defs要求是一个tuple/list，里面所有的元素是(column_name, colunm_width)
        '''
        wb = Workbook()
        ws = wb.active # insert at first position
        ws.title = sheet_name
        for i in range(len(columns)):
            c = ws.cell(row = 1, column = i+1)
            title, width = columns[i][0:2]
            if isinstance(title, str): title = title.decode('utf8')
            c.value = title
            c.style = Style(font=Font(bold=True))
            ws.column_dimensions[string.uppercase[i]].width = width

        row = 2
        for record in records:
            for i in range(len(record)):
                if record[i] == None:
                    continue
                c = ws.cell(row = row, column = i+1)
                cell_value = record[i].decode('utf8') if isinstance(record[i], str) else record[i]
                c.value = cell_value
            row = row + 1
        wb.save(out_file)


def export_to_excel(filename, fields, data_list):
    '''
    fields: (('Excel字段名'， 宽度， 'data_list中的key'), ('店名', 30, 'StoreName'), ...)
    data_list: {k1: v1, 'StoreName': 'xxxx', ...}
    filename: 生成的Excel名称
    '''
    keys = [f[2] for f in fields]
    columns = [(f[0], f[1]) for f in fields]
    records = [[d.get(k, None) for k in keys] for d in data_list]
    ExcelHelper.RecordsToExcel(records, columns, filename)


def create_mysql_conn(mysql_conf):
    '''
        mysql_conf = {
            'host': host, 'port': port, 'db': db,
            'user': user, 'passwd': passwd
        }
    '''
    conn = MySQLdb.connect(host=mysql_conf['host'],
                           port=mysql_conf['port'],
                           user=mysql_conf['user'],
                           passwd=mysql_conf['passwd'],
                           db=mysql_conf['db'],
                           charset='utf8',
                           cursorclass=DictCursor)
    conn.text_factory = str
    return conn


def select_data(mysql_conf, sql_str):
    '''
    sql_str 是SELEC语句
    '''
    desc = []
    results = []
    conn = None
    try:
        conn = create_mysql_conn(mysql_conf)
        cursor = conn.cursor()
        cursor.execute(sql_str)

        desc = [(i[0], 30, i[0]) for i in cursor.description]
        r = cursor.fetchall()
        for i in r:
            result = {}
            for k, v in i.items():
                result[k] = datetime.strftime(v, '%Y-%M-%d') \
                        if isinstance(v, datetime) else v
            results.append(result)

    except Exception, ex:
        traceback.print_exc()
        raise Exception(ex)

    finally:
        if conn != None:
            conn.close()

    return desc, results


def main():
    mysql_conf = MYSQL_CONF
    sql_str = SQL_STR
    filename = FILE_NAME

    desc, results = select_data(mysql_conf, sql_str)
    field_list = [(d[0], 20, d[0]) for d in desc]
    export_to_excel(filename=filename, fields=field_list, data_list=results)


if __name__ == '__main__':
    main()
